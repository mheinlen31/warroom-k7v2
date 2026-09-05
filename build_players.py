#!/usr/bin/env python3
"""Build js/players.js for the draft guide -- ESPN's full draftable pool with
the fields a value model needs: projected season points, auction value, ADP,
consensus rank, injury flag and a short outlook blurb.

Run:  python3 build_players.py     (prep.sh does this for you)
"""
import json, re, unicodedata, urllib.request
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parent
SEASON = 2026
URL = ("https://lm-api-reads.fantasy.espn.com/apis/v3/games/ffl/seasons/"
       f"{SEASON}/segments/0/leaguedefaults/3?view=kona_player_info")
FILTER = {"players": {"limit": 2000, "sortDraftRanks": {
    "sortPriority": 100, "sortAsc": True, "value": "PPR"}}}
POS = {1: "QB", 2: "RB", 3: "WR", 4: "TE", 5: "K", 16: "D/ST"}
PRO_TEAM = {1: "ATL", 2: "BUF", 3: "CHI", 4: "CIN", 5: "CLE", 6: "DAL", 7: "DEN",
            8: "DET", 9: "GB", 10: "TEN", 11: "IND", 12: "KC", 13: "LV", 14: "LAR",
            15: "MIA", 16: "MIN", 17: "NE", 18: "NO", 19: "NYG", 20: "NYJ", 21: "PHI",
            22: "ARI", 23: "PIT", 24: "LAC", 25: "SF", 26: "SEA", 27: "TB", 28: "WSH",
            29: "CAR", 30: "JAX", 33: "BAL", 34: "HOU"}
NFL_ABBR = {"cardinals": "ari", "falcons": "atl", "ravens": "bal", "bills": "buf",
            "panthers": "car", "bears": "chi", "bengals": "cin", "browns": "cle",
            "cowboys": "dal", "broncos": "den", "lions": "det", "packers": "gb",
            "texans": "hou", "colts": "ind", "jaguars": "jax", "chiefs": "kc",
            "raiders": "lv", "chargers": "lac", "rams": "lar", "dolphins": "mia",
            "vikings": "min", "patriots": "ne", "saints": "no", "giants": "nyg",
            "jets": "nyj", "eagles": "phi", "steelers": "pit", "49ers": "sf",
            "seahawks": "sea", "buccaneers": "tb", "titans": "ten", "commanders": "wsh"}
SUFFIXES = {"jr", "sr", "ii", "iii", "iv", "v"}
OUTLOOK_TOP = 260      # blurbs only for players who could plausibly matter


def norm(name):
    s = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode()
    s = re.sub(r"[^a-z ]", "", s.lower().replace(".", " ").replace("'", ""))
    return " ".join(p for p in s.split() if p not in SUFFIXES)


def img(name, pos, pid):
    if pos == "D/ST":
        ab = NFL_ABBR.get(norm(name))
        return f"https://a.espncdn.com/i/teamlogos/nfl/500/{ab}.png" if ab else None
    return f"https://a.espncdn.com/i/headshots/nfl/players/full/{pid}.png" if pid else None


# ------------------------------------------------------------ bye weeks
def fetch_bye_weeks():
    """{'BUF': 7, ...} for the season, from ESPN's pro-team schedule view."""
    url = f"https://lm-api-reads.fantasy.espn.com/apis/v3/games/ffl/seasons/{SEASON}?view=proTeamSchedules_wl"
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=60) as r:
        teams = json.load(r)["settings"]["proTeams"]
    return {t["abbrev"]: t.get("byeWeek") for t in teams if t.get("id") and t.get("byeWeek")}


# ------------------------------------------------------------ 2025 actuals
# ESPN stat ids, checked against 2025 totals (Gibbs 1,223 rush yds = id 24 ...)
STAT_IDS = {
    "QB":   {"py": 3, "ptd": 4, "int": 20, "ry": 24, "rtd": 25},
    "RB":   {"ra": 23, "ry": 24, "rtd": 25, "rec": 53, "recy": 42, "rectd": 43},
    "WR":   {"tgt": 58, "rec": 53, "recy": 42, "rectd": 43, "ry": 24, "rtd": 25},
    "TE":   {"tgt": 58, "rec": 53, "recy": 42, "rectd": 43},
    "K":    {"fgm": 83, "fga": 84, "xpm": 86},
    "D/ST": {"sack": 99, "int": 95, "fr": 96, "pa": 120, "ya": 127},
}
DST_TD_IDS = (101, 102, 103, 104)   # KR, PR, INT-return, FR-return touchdowns (106 is not a TD)


def last_season(p, pos):
    """2025 actuals: fantasy points under the league-default (PPR) scoring, games,
    and the handful of raw stats that matter for the position."""
    for s in p.get("stats") or []:
        if (s.get("statSourceId") == 0 and s.get("statSplitTypeId") == 0
                and s.get("seasonId") == SEASON - 1):
            raw = s.get("stats") or {}
            g = lambda i: raw.get(str(i), 0) or 0
            out = {k: round(g(i)) for k, i in STAT_IDS.get(pos, {}).items()}
            if pos == "D/ST":
                out["td"] = round(sum(g(i) for i in DST_TD_IDS))
            return {"fp25": round(float(s.get("appliedTotal") or 0), 1), "gp25": round(g(210)), "s25": out}
    return {"fp25": None, "gp25": None, "s25": None}


# ------------------------------------------------------------ experience
EXP_CACHE = ROOT / "data" / "experience.json"


def fetch_experience(ids):
    """{id: {'years': n, 'draftYear': y}} from ESPN's core athlete endpoint,
    cached on disk so a rebuild only fetches players it hasn't seen."""
    import concurrent.futures
    cache = {}
    if EXP_CACHE.exists():
        try:
            cache = json.loads(EXP_CACHE.read_text())
        except ValueError:
            cache = {}
    todo = [i for i in ids if i and str(i) not in cache]

    def one(pid):
        url = f"https://sports.core.api.espn.com/v2/sports/football/leagues/nfl/athletes/{pid}"
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=20) as r:
                a = json.load(r)
            return str(pid), {"years": (a.get("experience") or {}).get("years"),
                              "draftYear": (a.get("draft") or {}).get("year")}
        except Exception:
            return str(pid), {"years": None, "draftYear": None}

    if todo:
        with concurrent.futures.ThreadPoolExecutor(max_workers=16) as ex:
            for pid, rec in ex.map(one, todo):
                cache[pid] = rec
        EXP_CACHE.parent.mkdir(exist_ok=True)
        EXP_CACHE.write_text(json.dumps(cache))
        print(f"experience: fetched {len(todo)}, cached {len(cache)}")
    return cache


def experience_flags(rec):
    """rookie = entering year 1 (2026 draft class or an undrafted first-year);
    soph = entering year 2 (the 2025 class)."""
    if not rec:
        return False, False
    dy, yrs = rec.get("draftYear"), rec.get("years")
    if dy:
        return dy == SEASON, dy == SEASON - 1
    return yrs == 1, yrs == 2


def projection(p):
    """ESPN's projected season total: statSourceId 1 = projection,
    statSplitTypeId 0 = full season."""
    for s in p.get("stats") or []:
        if (s.get("statSourceId") == 1 and s.get("statSplitTypeId") == 0
                and s.get("seasonId") == SEASON):
            return round(float(s.get("appliedTotal") or 0), 1)
    return 0.0


# ------------------------------------------------------------------ sources
SOURCES = ROOT / "sources"
HDR = {
    "name": ["player name", "player", "name"],
    "pos": ["position", "pos"],
    "rank": ["overall rank", "overall", "ecr", "rank", "rk", "#"],
    "posrank": ["position rank", "pos rank", "posrank", "prk"],
    "proj": ["projected points", "fantasy points", "projection", "proj", "fpts", "fps", "points", "pts"],
    "aav": ["auction value", "avg $", "auction", "auc$", "auc", "aav", "value", "$"],
}
# outside sources' spellings -> ESPN's
SRC_ALIASES = {
    "kenneth gainwell": "kenny gainwell", "dermarcus robinson": "demarcus robinson",
    "braxton barrios": "braxton berrios", "laquan treadwell": "laquon treadwell",
}
POS_ALIAS = {"DST": "D/ST", "DEF": "D/ST", "D": "D/ST", "PK": "K"}


def _grid(path):
    """every row of the file as a list of cell strings"""
    if path.suffix.lower() == ".csv":
        import csv
        with open(path, newline="", encoding="utf-8-sig") as fh:
            return [[(c or "").strip() for c in r] for r in csv.reader(fh)]
    from openpyxl import load_workbook
    ws = load_workbook(path, read_only=True, data_only=True).active
    return [["" if v is None else str(v).strip() for v in row] for row in ws.iter_rows(values_only=True)]


def _tables(path):
    """yield (positional, rows) for each table in the file.

    A plain export is one table. A POSITIONAL GUIDE lays several tables side
    by side -- one RK/Player/FPS/... block per position, separated by blank
    columns (The Athletic's sheet is built this way). Each block becomes its
    own table, flagged positional so its RK is read as the rank within the
    position rather than an overall rank."""
    grid = _grid(path)
    if not grid:
        return
    hdr = [h.lower() for h in grid[0]]
    blocks, cur = [], []
    for i, h in enumerate(hdr):
        if h == "":
            if cur:
                blocks.append(cur); cur = []
        else:
            cur.append(i)
    if cur:
        blocks.append(cur)
    blocks = [bk for bk in blocks if any(hdr[i] in HDR["name"] for i in bk)]
    positional = len(blocks) > 1
    for cols in blocks:
        rows = []
        for row in grid[1:]:
            rows.append({hdr[i]: (row[i] if i < len(row) else "") for i in cols})
        yield positional, rows


def _col(row, key):
    for h in HDR[key]:
        if h in row and row[h] != "":
            return row[h]
    return None


def _num(v):
    try:
        return float(str(v).replace("$", "").replace(",", ""))
    except (TypeError, ValueError):
        return None


def _clean_name(raw, pos):
    n = re.sub(r"\(.*?\)", "", raw)                  # "Josh Allen (BUF)"
    n = re.sub(r"\b(D/ST|DST|Defense)\b", "", n, flags=re.I).strip(" ,-")
    if pos == "D/ST" and n:
        n = n.split()[-1]                             # "Baltimore Ravens" -> Ravens
    return n


def _config():
    """sources/sources.json (optional): per-file label, weight, and whether its
    auction values are MARKET prices (blend into Mkt $) or a model's values
    (keep out of Mkt $ -- they still feed Model $ through projections)."""
    cfg = SOURCES / "sources.json"
    if not cfg.exists():
        return {}
    try:
        return json.loads(cfg.read_text())
    except ValueError as e:
        print(f"sources.json is not valid JSON ({e}); ignoring it")
        return {}


def read_sources():
    """-> list of {name, label, rows{norm_name: {pos, rank, posrank, proj, aav, raw}}, aav, weight}"""
    out = []
    if not SOURCES.exists():
        return out
    cfg = _config()
    for path in sorted(SOURCES.iterdir()):
        if path.name.startswith(("_", ".")) or path.suffix.lower() not in (".csv", ".xlsx", ".xlsm"):
            continue
        c = cfg.get(path.stem, {})
        label = c.get("label") or "".join(w[0] for w in path.stem.split() if w[0].isalnum())[:4].upper()
        rows, bad, tables = {}, 0, 0
        for positional, table in _tables(path):
            tables += 1
            for r in table:
                raw = _col(r, "name")
                if not raw:
                    continue
                pos_raw = (_col(r, "pos") or "").upper()
                m = re.match(r"([A-Z/]+)(\d+)?", pos_raw)
                pos = POS_ALIAS.get(m.group(1), m.group(1)) if m else None
                rank = _num(_col(r, "rank"))
                posrank = _num(_col(r, "posrank")) or (float(m.group(2)) if m and m.group(2) else None)
                if positional and rank is not None and posrank is None:
                    posrank, rank = rank, None       # per-position table: RK is the rank within the position
                name = _clean_name(raw, pos)
                if not name:
                    bad += 1
                    continue
                key = norm(name); key = SRC_ALIASES.get(key, key)
                rows[key] = {"pos": pos, "rank": rank, "posrank": posrank,
                                    "proj": _num(_col(r, "proj")), "aav": _num(_col(r, "aav")), "raw": raw}
        out.append({"name": path.stem, "label": label, "rows": rows,
                    "aav": bool(c.get("aav", True)), "weight": float(c.get("weight", 1) or 1)})
        print(f"source {path.name} [{label}]: {len(rows)} rows"
              + (f" across {tables} position tables" if tables > 1 else "")
              + (f", {bad} unreadable" if bad else ""))
    return out


def _wmean(pairs):
    w = sum(x[1] for x in pairs)
    return sum(v * x for v, x in pairs) / w if w else 0


def blend(players, sources):
    """Fold outside sources into the ESPN pool. Consensus rank is the weighted
    mean overall rank across everything that has one; projection is the
    weighted mean of ESPN's and each source's (a rank-only source gets an
    implied projection: ESPN's projection at that positional rank); auction
    value likewise, but only from sources flagged as market prices."""
    for p in players:
        p["projEspn"] = p["proj"]
        p["srcRanks"], p["srcPos"], p["srcProj"] = {}, {}, {}
        p["nsrc"] = 1
    if not sources:
        return
    by_norm = {norm(p["name"]): p for p in players}
    dst = {norm(p["name"]): p for p in players if p["pos"] == "D/ST"}
    curve = {}
    for pos in set(p["pos"] for p in players):
        curve[pos] = [p["proj"] for p in sorted((x for x in players if x["pos"] == pos),
                                                key=lambda x: -x["proj"]) if p["proj"] > 0]
    for src in sources:
        label, rows, w = src["label"], src["rows"], src["weight"]
        matched, unmatched = [], []
        for k, r in rows.items():
            p = by_norm.get(k) or (dst.get(k.split()[-1]) if k else None)   # "Houston Texans" -> Texans
            if not p:
                unmatched.append(r["raw"])
                continue
            if not r["pos"]:
                r["pos"] = p["pos"]            # a positional guide carries no POS column
            matched.append((k, r, p))
        by_pos = {}
        for k, r, p in matched:
            if r["rank"] is not None:
                by_pos.setdefault(r["pos"], []).append((r["rank"], k))
        implied_prk = {}
        for pos, lst in by_pos.items():
            for i, (_, k) in enumerate(sorted(lst)):
                implied_prk[k] = i + 1
        # Put the source's projections on ESPN's scale, position by position:
        # a 6-point passing TD or a different yardage rate shifts a whole
        # position, and we want the source's ORDER and GAPS, not its scoring
        # settings. Ratio of summed projections over the players both have.
        scale = {}
        for pos in set(r["pos"] for _, r, _ in matched):
            both = [(p["projEspn"], r["proj"]) for _, r, p in matched
                    if r["pos"] == pos and r["proj"] and r["proj"] > 0 and p["projEspn"] > 0]
            both.sort(key=lambda t: -t[0])
            both = both[:24]
            if len(both) >= 5:
                scale[pos] = min(1.3, max(0.7, sum(e for e, _ in both) / sum(s for _, s in both)))
        if scale:
            print(f"  {label} projection scale to ESPN: " + " ".join(f"{k} {v:.2f}" for k, v in sorted(scale.items())))
        for k, r, p in matched:
            if r["proj"] and r["proj"] > 0 and r["pos"] in scale:
                r["proj"] = r["proj"] * scale[r["pos"]]
            projs = p.setdefault("_projs", [(p["projEspn"], 1.0)] if p["projEspn"] > 0 else [])
            aavs = p.setdefault("_aavs", [(p["aav"], 1.0)] if p["aav"] > 0 else [])
            ranks = p.setdefault("_ranks", ([(p["cons"], 1.0)] if p.get("cons") else []))
            used = False
            if r["rank"] is not None:
                ranks.append((r["rank"], w)); p["srcRanks"][label] = r["rank"]; used = True
            prk = r["posrank"] or implied_prk.get(k)
            if prk:
                p["srcPos"][label] = int(prk); used = True
            if r["proj"] is not None and r["proj"] > 0:
                projs.append((r["proj"], w)); p["srcProj"][label] = round(r["proj"], 1); used = True
            elif prk:
                cv = curve.get(p["pos"], [])
                if 1 <= int(prk) <= len(cv):
                    projs.append((cv[int(prk) - 1], w)); used = True
            if src["aav"] and r["aav"] is not None and r["aav"] > 0:
                aavs.append((r["aav"], w)); used = True
            if used:
                p["_nsrc"] = p.get("_nsrc", 0) + 1
        print(f"  {label}: matched {len(matched)} of {len(rows)}"
              + (f"; not in the ESPN pool: {', '.join(unmatched[:20])}{' …' if len(unmatched) > 20 else ''}" if unmatched else ""))
    for p in players:
        if len(p.get("_projs", [])) > 1:
            p["proj"] = round(_wmean(p["_projs"]), 1)
        if len(p.get("_aavs", [])) > 1:
            p["aav"] = round(_wmean(p["_aavs"]), 1)
        if len(p.get("_ranks", [])) > 1:
            p["cons"] = round(_wmean(p["_ranks"]), 1)
            p["spread"] = round(max(v for v, _ in p["_ranks"]) - min(v for v, _ in p["_ranks"]))
        p["nsrc"] = 1 + p.pop("_nsrc", 0)
        for k in ("_projs", "_aavs", "_ranks"):
            p.pop(k, None)
    print(f"blended {len(sources)} outside source(s) into the pool")


def main():
    req = urllib.request.Request(URL, headers={
        "X-Fantasy-Filter": json.dumps(FILTER), "User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=90) as r:
        data = json.load(r)

    seen, out, rankings_shape = set(), [], None
    for entry in data.get("players", []):
        p = entry.get("player") or {}
        pos = POS.get(p.get("defaultPositionId"))
        if not pos:
            continue
        name = (p.get("fullName") or "").replace(" D/ST", "").strip()
        if not name or norm(name) in seen:
            continue
        seen.add(norm(name))
        own = p.get("ownership") or {}
        ranks = p.get("draftRanksByRankType") or {}
        rank = ((ranks.get("PPR") or ranks.get("STANDARD") or {}).get("rank")) or 9999
        if rankings_shape is None and p.get("rankings"):
            rankings_shape = json.dumps(p["rankings"])[:300]
        inj = p.get("injuryStatus")
        # ESPN publishes several rankers per player; the mean PPR rank across
        # them is a real consensus, and the spread says how much they disagree
        ppr = [r["rank"] for grp in (p.get("rankings") or {}).values()
               for r in (grp or []) if r.get("rankType") == "PPR" and r.get("rank")]
        cons = round(sum(ppr) / len(ppr), 1) if ppr else None
        spread = (max(ppr) - min(ppr)) if len(ppr) > 1 else 0
        rec = {
            "id": p.get("id"),
            "name": name, "pos": pos,
            "nfl": PRO_TEAM.get(p.get("proTeamId")) if pos != "D/ST" else None,
            "img": img(name, pos, p.get("id")),
            "aav": round(own.get("auctionValueAverage") or 0, 1),
            "adp": round(own.get("averageDraftPosition") or 0, 1),
            "rank": rank,
            "proj": projection(p),
            "inj": None if not inj or inj == "ACTIVE" else inj,
            "cons": cons, "spread": spread, "nrank": len(ppr),
            **last_season(p, pos),
        }
        if rank <= OUTLOOK_TOP and p.get("seasonOutlook"):
            o = re.sub(r"\s+", " ", p["seasonOutlook"]).strip()
            rec["outlook"] = (o[:300] + "…") if len(o) > 300 else o
        out.append(rec)

    # bye weeks (D/ST rows carry the team nickname; map it to the abbreviation)
    byes = fetch_bye_weeks()
    for rec in out:
        ab = rec["nfl"] or (NFL_ABBR.get(norm(rec["name"]), "").upper() if rec["pos"] == "D/ST" else None)
        rec["bye"] = byes.get(ab) if ab else None
    print("bye weeks:", sum(1 for r in out if r["bye"]), "of", len(out))
    # experience -> rookie / second-year flags
    expc = fetch_experience([r["id"] for r in out if r["pos"] != "D/ST"])
    for rec in out:
        rk, so = experience_flags(expc.get(str(rec["id"])))
        rec["rookie"], rec["soph"] = rk, so
        rec.pop("id", None)
    print("rookies:", sum(1 for r in out if r["rookie"]), "| second-year:", sum(1 for r in out if r["soph"]))
    blend(out, read_sources())
    out.sort(key=lambda x: (x["rank"], -x["aav"], x["name"]))
    payload = {"season": SEASON,
               "generated": datetime.now(timezone.utc).isoformat(timespec="seconds"),
               "players": out}
    (ROOT / "js" / "players.js").write_text(
        "window.GUIDE_PLAYERS = " + json.dumps(payload) + ";\n")
    print(f"wrote {len(out)} players:", dict(Counter(p['pos'] for p in out)))
    print("with projection:", sum(1 for p in out if p["proj"] > 0),
          "| with aav:", sum(1 for p in out if p["aav"] > 0),
          "| injured flagged:", sum(1 for p in out if p["inj"]))
    print("with consensus:", sum(1 for p in out if p["cons"]),
          "| rankers seen up to:", max((p["nrank"] for p in out), default=0))


if __name__ == "__main__":
    main()
